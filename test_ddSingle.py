import requests
import json
import jsonpath
import openpyxl


def test_addMultipleStudent():
    #set url and read request json file

    url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open("newStudent")
    json_request = json.loads(f.read())

    #reading excel
    workbook = openpyxl.load_workbook('ddTesting.xlsx')
    sh = workbook['Sheet1']
    rows = sh.max_row       #to calculate how many records we have - 1 header

    for i in range(2,rows+1):           #last record is skipped hence rows+1
        cell_first_name = sh.cell(row=i, column=1)          #to iterate record for each row as column is fixed
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)
#        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value
        response = requests.post(url,json_request)
        print(response.text)
        print(response.status_code)
        assert(response.status_code == 201)
        print("Executed")
