import json
import jsonpath
import openpyxl
import requests

class Common:

    def __init__(self, FilePathName, SheetName):
        global workbook
        global sh

        workbook1 = openpyxl.load_workbook(FilePathName)
        sh = workbook1[SheetName]


    def fetch_row_count(self):
        # reading excel
        rows = sh.max_row   # to calculate how many records we have - 1 header
        return rows

    def fetch_Column_Count(self):
        column = sh.max_column
        return column


    def fetch_keyNames(self):
        columnCount = sh.max_column
        keys = []
        for i in range(1, columnCount+1):
            cell = sh.cell(row=1, column=i)
            keys.insert(i-1, cell.value)
        return keys

    def update_json_request_with_data(self, rowNumber, jsonRequest, keyList):
        colCount = sh.max_column

        for i in range(1, colCount+1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest
